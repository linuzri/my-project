""" read config from config.xls file, execute command & save in
    different colum in same excel file """

import openpyxl
from netmiko import ConnectHandler
import time

cisco1 = {
    'device_type': 'cisco_ios',
    'host': '192.168.137.101',
    'username': 'admin',
    'password': 'cisco',
}
cisco2 = {
    'device_type': 'cisco_ios',
    'host': '192.168.137.102',
    'username': 'admin',
    'password': 'cisco',
}
cisco3 = {
    'device_type': 'cisco_ios',
    'host': '192.168.137.103',
    'username': 'admin',
    'password': 'cisco',
}

for i in range(1,4):
    excel_file = openpyxl.load_workbook("config.xlsx")
    device_config = excel_file['SW'+str(i)]

    for config in range(2, device_config.max_row + 1):
        cmd = device_config.cell(config, 2).value
        with open(cisco1['host'] + '.txt', 'a') as file:
            file.write(cmd + "\n")
        # print(file)

# command = "show ip int brief"
# fileconfig = file.read()

# with open(cisco1['host'] + '.txt') as file:
# with ConnectHandler(**cisco1) as net_connect:
   # output = net_connect.send_command(file.read())
   # print(output)

# for devices in (cisco1,cisco2,cisco3):
#    cfg_file = devices['host'] + '.txt'
#    net_connect = ConnectHandler(**devices)
#    print("connecting to " + devices['host'] + "...")
#    time.sleep(1)
#    net_connect.send_config_from_file(cfg_file)
#    print("pushing config for " + devices['host'] + "...")
#    time.sleep(1)
#    net_connect.save_config()
#    print("saving config for " + devices['host'] + "...")
#    time.sleep(1)
#    net_connect.disconnect()
#    print("disconnecting from " + devices['host'] + "...")
#    time.sleep(1)
#    print("done!")

# with ConnectHandler(**cisco1) as net_connect:
#   with open('SW1.txt') as file:
#      fileconfig = file.read()
#      output = net_connect.send_command(fileconfig)
#     print(output)
